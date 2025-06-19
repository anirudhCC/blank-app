import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
import pandas as pd
from fpdf import FPDF
import os
import zipfile
import re

def load_excel(file_path):
    return load_workbook(filename=file_path, data_only=True)

def sanitize_filename(name):
    name = re.sub(r'[<>:"/\\|?*]', '', str(name))
    name = name.replace(' ', '_')[:50]
    return name or "default"

def get_page_ranges(sheet):
    horizontal_breaks = [0] + [page_break.id for page_break in sheet.row_breaks.brk]
    last_cell = sheet.dimensions.split(":")[1]
    last_col = ''.join(c for c in last_cell if c.isalpha())
    last_row = ''.join(c for c in last_cell if c.isdigit())
    pages = {}
    for idx, h_break in enumerate(horizontal_breaks):
        start_row = h_break + 1
        end_row = horizontal_breaks[idx + 1] if idx + 1 < len(horizontal_breaks) else int(last_row)
        if idx == 0: start_row = 2
        pages[f"Page{idx + 1}"] = f"A{start_row}:{last_col}{end_row}"
    return pages

def get_table_styles(sheet):
    styles = {}
    for cell in sheet[1]:
        font_name = cell.font.name.lower() if cell.font.name else 'arial'
        styles[cell.column_letter] = {'font': font_name, 'bold': cell.font.bold}
    return styles

def calculate_column_widths(headers, data, pdf, max_width=500, fixed_width=30):
    col_widths = []
    for j, header in enumerate(headers):  #CHANGED: Added header to check for "Details"
        pdf.set_font('Arial', 'B', 8)
        width = pdf.get_string_width(str(header) or "") + 6  # CHANGED: Adjusted padding to 6
        if header == "Details":  # CHANGED: Increase width for "Details" column
            width = max(width, 50)  # CHANGED: Minimum width of 60 for "Details"
        else:
            width = min(width, 30)  # CHANGED: Cap other columns at 30
        col_widths.append(width)
    total_width = sum(col_widths)
    if total_width > max_width:
        col_widths = [w * (max_width / total_width) for w in col_widths]
    return col_widths

def save_as_pdf(sheet, page_ranges, output_folder):
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    pdf_files = []
    headers = [cell.value for cell in sheet[1]]
    styles = get_table_styles(sheet)

    for idx, cell_range in page_ranges.items():
        data = pd.DataFrame([[cell.value for cell in row] for row in sheet[cell_range]])
        brand_idx = 2
        #headers.index('BrandSupplierDescription') if 'BrandSupplierDescription' in headers else headers.index('Supplier') if 'Supplier' in headers else None       

        ccn_idx = headers.index('CCN') if 'CCN' in headers else None
        if data.empty:
            print("Data is unexpectedly empty before file name creation.")
            print(cell_range)
            file_name = f"{idx}.pdf"
        else:
            file_name = (f"{sanitize_filename(data.iloc[0, brand_idx])}_{sanitize_filename(data.iloc[0, ccn_idx])}.pdf"
                            if brand_idx is not None and ccn_idx is not None 
                            else f"{idx}.pdf"
                            )

        #file_name = f"{sanitize_filename(data.iloc[0, brand_idx])}_{sanitize_filename(data.iloc[0, ccn_idx])}.pdf" \
                    #if brand_idx is not None and ccn_idx is not None else f"{idx}.pdf"

        pdf_file_path = os.path.join(output_folder, file_name)
        pdf = FPDF(orientation='L')
        pdf.add_page()
        pdf.set_font('Arial', '', 8)
        row_height = 4
        col_widths = calculate_column_widths(headers, data, pdf)

        # Headers
        start_y = pdf.get_y()
        start_x = pdf.l_margin  # Start at the left margin
        pdf.set_xy(start_x, start_y)
        for j, header in enumerate(headers):
            col_letter = get_column_letter(j + 1)
            if header in ['Contract_Reference_Number', 'BrandSupplierDescription']:
                pdf.set_font('Arial', 'B', 6)
            else:
                pdf.set_font('Arial', 'B', 8)
            pdf.set_fill_color(200, 200, 200)
            pdf.cell(col_widths[j], row_height * 2, str(header) or "", border=1, fill=True, align='C')
            start_x += col_widths[j]  # Move to the next column position
            pdf.set_xy(start_x, start_y)
        pdf.ln(row_height * 2)


        # Data
        effective_page_height = pdf.h - pdf.t_margin - pdf.b_margin  # Usable height (e.g., ~170 for landscape A4)
        current_y = pdf.get_y()
        for row_idx, row in data.iterrows():
            if current_y + row_height > effective_page_height:
                pdf.add_page()
                current_y = pdf.get_y()

            start_y = current_y
            max_height = row_height
            start_x = pdf.l_margin  # Reset X to the left margin for each row
            pdf.set_xy(start_x, start_y)
            row_color = (255, 255, 255) if row_idx % 2 == 0 else (230, 230, 230)
            pdf.set_fill_color(*row_color)


            for j, item in enumerate(row):
                col_letter = get_column_letter(j + 1)
                if headers[j] in ['Item Name']:
                    pdf.set_font('Arial', '', 6)
                else:
                    pdf.set_font('Arial', '', 7)
                if headers[j] in ['Contract', 'Contract_Reference_Number']:
                    if pd.isna(item) or item is None:
                         item = ""  
                    else:
                        item = int(item)

                if isinstance(item, (int, float)):
                    item = round(item, 2)
               
                if headers[j] in ['RetroRate','Rate','Volume', 'Retro_Value', 'Retro Rate'] and item is not None and not pd.isna(item):
                    content = f"£{item:.2f}"
                else:
                    content="" if pd.isna(item) or item is None else str(item)
                

                #if isinstance(item, (int, float)):
                    #item = round(item, 2)
                #content = "" if pd.isna(item) or item is None else str(item)
                #if headers[j] in ['RetroRate', 'Retro_Value','Retro Rate'] and content:
                    #content = f"£{content}"
                
                lines = []
                words = content.split(' ')
                current_line = ""
                for word in words:
                    if pdf.get_string_width(current_line + word) < col_widths[j]:
                        current_line += word + " "
                    else:
                        lines.append(current_line.strip())
                        current_line = word + " "
                lines.append(current_line.strip())
                wrapped_content = '\n'.join(lines)
                
                pdf.multi_cell(col_widths[j], row_height, wrapped_content, align='C', border=0,fill=True)
                cell_height = pdf.get_y() - start_y
                max_height = max(max_height, cell_height)
                start_x += col_widths[j]  # Move to the next column position
                pdf.set_xy(start_x, start_y)  # Set position for the next cell

            current_y = start_y + max_height
            pdf.set_xy(pdf.l_margin, current_y)
            #print(f"Row processed: start_y={start_y}, max_height={max_height}, current_page={pdf.page_no()}")

        pdf.output(pdf_file_path)
        print(f"PDF {file_name} generated with {pdf.page_no()} pages")
        pdf_files.append(pdf_file_path)
    
    return pdf_files

def zip_pdfs(pdf_files, zip_filename="generated_pdfs.zip"):
    with zipfile.ZipFile(zip_filename, 'w') as zipf:
        for pdf_file in pdf_files:
            zipf.write(pdf_file)
    return zip_filename

def main():
    st.title("Excel to PDF Generator")
    uploaded_file = st.file_uploader("Upload your Excel file", type=["xlsx"])

    if uploaded_file is not None:
        file_path = uploaded_file.name
        with open(file_path, "wb") as f:
            f.write(uploaded_file.getbuffer())

        try:
            workbook = load_excel(file_path)
            sheet = workbook.worksheets[0]
            page_ranges = get_page_ranges(sheet)

            if not page_ranges:
                st.warning("⚠️ No page breaks detected. Ensure page breaks are set.")
                return

            pdf_files = save_as_pdf(sheet, page_ranges, "generated_pdfs")
            zip_file = zip_pdfs(pdf_files)
            
            with open(zip_file, "rb") as file:
                st.download_button(
                    label="📥 Download PDFs",
                    data=file,
                    file_name=zip_file,
                    mime="application/zip"
                )
            
            st.success(f"{len(pdf_files)} PDFs successfully generated!")
            

        except FileNotFoundError as e:
            st.error(f"Error: {e}")
        except Exception as e:
            st.error(f"Unexpected error: {e}")

if __name__ == "__main__":
    main()
